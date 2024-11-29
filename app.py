
import requests
import json
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# API keys for SerpAPI and GroQ
SERPAPI_KEY = "5db04aca1cb21545567c03365887bc5208a1e3b022f029d695189a62e5b449c2"
GROQ_API_KEY = "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"

def fetch_datasets_with_serpapi(keywords):
    """
    Use SerpAPI to search for dataset links from Kaggle, GitHub, and HuggingFace.
    Returns a dictionary with platform names as keys and clickable links as values.
    """
    queries = {
        "Kaggle": f"{keywords} site:kaggle.com",
        "GitHub": f"{keywords} site:github.com",
        "HuggingFace": f"{keywords} site:huggingface.co"
    }

    dataset_links = {}

    for platform, query in queries.items():
        search_url = "https://serpapi.com/search"
        params = {
            "q": query,
            "api_key": SERPAPI_KEY,
            "num": 5  # Fetch top 5 results
        }

        try:
            response = requests.get(search_url, params=params)
            if response.status_code == 200:
                results = response.json().get("organic_results", [])
                links = [result["link"] for result in results if "link" in result]
                dataset_links[platform] = links
            else:
                dataset_links[platform] = [f"Error fetching datasets from {platform} (status code: {response.status_code})"]
        except Exception as e:
            dataset_links[platform] = [f"Error fetching datasets from {platform}: {e}"]

    return dataset_links

def search_with_serpapi(query):
    """
    Perform a web search using SerpAPI.
    """
    params = {
        "q": query,
        "api_key": SERPAPI_KEY,
        "num": 2  # Number of results to fetch
    }
    url = "https://serpapi.com/search"
    response = requests.get(url, params=params)
    if response.status_code == 200:
        return response.json().get("organic_results", [])
    else:
        return {"error": f"Error: {response.status_code}, {response.text}"}

def research_industry(company_name):
    """
    Perform market research on the company and its industry.
    """
    queries = [
        f"{company_name} industry analysis",
        f"{company_name} business model",
        f"{company_name} key offerings",
        f"{company_name} strategic focus areas",
        f"{company_name} AI adoption trends",
        f"{company_name} market trends",
        f"{company_name} AI and automation adoption",
        f"{company_name} supply chain strategy",
        f"{company_name} customer experience strategy",
        f"{company_name} vision and product innovations"
    ]
    
    results = {query: search_with_serpapi(query) for query in queries}
    return results

def extract_snippets(data):
    """
    Extract the 'snippet' field from the search results.
    """
    snippets = {}
    for query, results in data.items():
        snippets[query] = [result.get("snippet", "No snippet available") for result in results]
    return snippets

def save_to_file(data, file_name):
    """
    Save data to a JSON file.
    """
    with open(file_name, "w") as file:
        json.dump(data, file, indent=4)
    print(f"Data saved to {file_name}")

def generate_use_cases_with_groq(snippets):
    """
    Use GroQ API to generate AI/ML use cases based on the provided snippets.
    """
    combined_snippets = ""
    for key, values in snippets.items():
        combined_snippets += f"{key}:\n" + "\n".join(values) + "\n\n"
    
    payload = {
        "model": "llama3-8b-8192",
        "messages": [
            {
                "role": "user",
                "content": f"Based on the following company and industry information, generate 5 actionable use cases where the company can leverage Generative AI (GenAI), Large Language Models (LLMs), and Machine Learning (ML) technologies to improve their processes, enhance customer satisfaction, and boost operational efficiency. For each use case, provide a title and a detailed description of how the technology would be applied:\n\n{combined_snippets}"
            }
        ],
        "max_tokens": 1024,
        "temperature": 1
    }

    api_url = "https://api.groq.com/openai/v1/chat/completions"
    headers = {
        "Authorization": f"Bearer {GROQ_API_KEY}",
        "Content-Type": "application/json"
    }

    response = requests.post(api_url, headers=headers, data=json.dumps(payload))
    
    if response.status_code == 200:
        return response.json().get("choices", [])
    else:
        return {"error": f"Error: {response.status_code}, {response.text}"}

def generate_datasets_for_use_cases(use_cases):
    """
    For each use case title, fetch datasets using the SerpAPI.
    """
    datasets = {}
    for use_case in use_cases:
        if "message" in use_case and "content" in use_case["message"]:
            content = use_case["message"]["content"]
            lines = content.split("\n")
            titles = [line.split(":")[1].strip() for line in lines if line.startswith("Title")]
            # print(titles)
            
            for title in titles:
                # Fetch dataset links for each title
                dataset_links = fetch_datasets_with_serpapi(title)
                datasets[title] = dataset_links
    
    return datasets

def save_datasets_to_file(datasets, file_name):
    """
    Save datasets to a JSON file.
    """
    with open(file_name, "w") as file:
        json.dump(datasets, file, indent=4)
    print(f"Dataset links saved to {file_name}")

def save_use_cases_to_excel_with_links(use_cases, datasets_file, file_name):
    """
    Save the generated use cases to an Excel sheet with three columns: Use Case, Description, Dataset Links.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Use Cases"

    # Add headers
    ws.append(["Use Case", "Description", "Dataset Links"])

    # Read dataset links from the 'datasets.json' file
    with open(datasets_file, "r") as file:
        dataset_links = json.load(file)

    # Add use case data
    for use_case in use_cases:
        if "message" in use_case and "content" in use_case["message"]:
            content = use_case["message"]["content"]
            lines = content.split("\n")
            titles = [line.split(":")[1].strip() for line in lines if line.startswith("Title")]
            descriptions = [line.split(":")[1].strip() for line in lines if line.startswith("Description")]

            for title, description in zip(titles, descriptions):
                # Get dataset links for the title
                platform_links = []
                for platform in dataset_links:
                    platform_links.extend(dataset_links[platform])

                # Create the clickable dataset links string
                clickable_links = "/n".join([f'=HYPERLINK("{link}", "{link}")' for link in platform_links])

                # Add the title, description, and dataset links as a hyperlink
                ws.append([title, description, clickable_links])

    # Auto-adjust the column width to fit the content
    for col in range(1, 4):
        column_letter = get_column_letter(col)
        max_length = 0
        for row in ws.iter_rows():
            cell_value = str(row[col - 1].value)
            max_length = max(max_length, len(cell_value))
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save the Excel file
    wb.save(file_name)
    print(f"Use cases with dataset links saved to Excel: {file_name}")
# Function Definition
def save_use_cases_to_excel_with_links(use_cases, datasets_file, file_name):
    """
    Save the generated use cases to an Excel sheet with three columns: Use Case, Description, Dataset Links.
    Each dataset link will be clickable and placed in the 'Dataset Links' column in separate rows.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Use Cases"

    # Add headers
    ws.append(["Use Case", "Description", "Dataset Links"])

    # Use the datasets_file_content directly instead of opening a file
    with open(datasets_file, "r") as file:
        dataset_links = json.load(file)

    # Add use case data
    for use_case in use_cases:
        if "message" in use_case and "content" in use_case["message"]:
            content = use_case["message"]["content"]
            lines = content.split("\n")
            titles = [line.split(":")[1].strip() for line in lines if line.startswith("Title")]
            descriptions = [line.split(":")[1].strip() for line in lines if line.startswith("Description")]

            for title, description in zip(titles, descriptions):
                # Get dataset links for the title
                platform_links = []
                for platform in dataset_links:
                    platform_links.extend(dataset_links[platform])

                # Add the use case title and description in one row
                row = [title, description]  # Use Case and Description in the first two columns
                ws.append(row)
                row_num = ws.max_row  # Get the last row number

                # Add each link in the next rows under "Dataset Links" for the current use case
                for link in platform_links:
                    # Ensure the link starts with "http://"
                    if not link.startswith(("http://", "https://")):
                        link = f"http://{link}"

                    # Move to the next row and add the hyperlink
                    ws.append([None, None, link])  # Leave Use Case and Description as None
                    new_row_num = ws.max_row  # Get the new row number

                    # Add the hyperlink to the 'Dataset Links' column
                    ws.cell(row=new_row_num, column=3, value=link).hyperlink = link  # Set the hyperlink property
                    ws.cell(row=new_row_num, column=3).style = "Hyperlink"  # Optional: Apply a hyperlink style

    # Auto-adjust the column width to fit the content
    for col in range(1, 4):
        column_letter = get_column_letter(col)
        max_length = 0
        for row in ws.iter_rows():
            cell_value = str(row[col - 1].value) if row[col - 1].value else ""
            max_length = max(max_length, len(cell_value))
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save the Excel file
    wb.save(file_name)
    print(f"Use cases with dataset links saved to Excel: {file_name}")
def main(company_name, output_path):
    """
    Main function to conduct industry research and generate use cases.
    """
    # Perform research
    print(f"Conducting research for {company_name}...")
    research_data = research_industry(company_name)
    
    # Save raw research data
    save_to_file(research_data, f"{output_path}/{company_name}_responses.json")
    
    # Extract snippets
    snippets = extract_snippets(research_data)
    snippets_file = f"{output_path}/{company_name}_snippets.json"
    save_to_file(snippets, snippets_file)
    
    # Generate use cases using the snippets
    print(f"Generating use cases for {company_name}...")
    use_cases = generate_use_cases_with_groq(snippets)
    
    # Generate dataset links for use cases and save to file
    datasets = generate_datasets_for_use_cases(use_cases)
    datasets_file = f"{output_path}/{company_name}_datasets.json"
    save_datasets_to_file(datasets, datasets_file)

    # Save use cases with dataset links to Excel
    use_cases_excel = f"{output_path}/{company_name}_use_cases_with_links.xlsx"
    save_use_cases_to_excel_with_links(use_cases, datasets_file, use_cases_excel)
    
    print(f"Snippets saved to: {snippets_file}")
    print(f"Dataset links saved to: {datasets_file}")
    print(f"Use cases with dataset links saved to Excel: {use_cases_excel}")


# Example Execution
if __name__ == "__main__":
    company_name = "infosys"
    output_path = r"C:\Users\user\Desktop\internship Project"
    main(company_name, output_path)
